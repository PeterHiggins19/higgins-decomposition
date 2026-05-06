#!/usr/bin/env python3
"""
Two more sibling tests for the Hs-05 within-domain robustness battery.

(a) Tappe (2024) — Kimberlite Group-1 bulk rocks (D=10 oxides),
    binned by Country/Region. A different intra-cratonic mantle-derived
    ultrapotassic rock type than Ball/Stracke. K2O is typically VERY
    high in kimberlites (>3%), so the K2O-prefix prediction is strongest
    here.

(b) Qin (2024) — Clinopyroxene mineral spot analyses from intra-cratonic
    mantle xenoliths and ultramafic rocks (D=9 oxides). K2O is NOT in
    the carrier set — Cr2O3 replaces it. Crucial test: does the
    "K2O-prefix" become an "alkali-prefix" when K2O is absent?
    If lineage starts with Na2O instead, the prefix is "dominant alkali"
    in general, not specifically potassium.

Outputs (in this folder):
  tappe_kim1_by_country_barycenters.csv     T=8, D=10
  qin_cpx_by_location_barycenters.csv       T=top-N, D=9
  *_summary.json
"""
import openpyxl, csv, json, math, collections, statistics
from pathlib import Path

ROOT = Path(__file__).resolve().parents[5]
OUT  = Path(__file__).resolve().parent
MIN_N = 10


def aitchison_barycenter(rows):
    if not rows: return None
    D = len(rows[0])
    log_means = [sum(math.log(r[j]) for r in rows) / len(rows) for j in range(D)]
    g = [math.exp(lm) for lm in log_means]
    return [x/sum(g) for x in g]


# ── (a) Tappe v2024 KIM1 ──
def bin_tappe_kim1():
    src = ROOT / "DATA" / "Geochemistry" / "2022-2-FLV19S_Tappe_data_v2024.xlsx"
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb["Major Elements Bulk KIM1"]

    # Header at L5: cols 0=Country/Region, 6=Age, 7..17 = oxides
    # Oxide order (out): SiO2 TiO2 Al2O3 FeO CaO MgO MnO K2O Na2O P2O5
    OXIDES = ["SiO2","TiO2","Al2O3","FeO","CaO","MgO","MnO","K2O","Na2O","P2O5"]
    oxide_cols = [7, 8, 9, 11, 14, 13, 12, 16, 15, 17]
    country_idx = 0

    by_country = collections.defaultdict(list)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= 5: continue
        if row is None or len(row) < 18: continue
        try:
            vals = [float(row[k]) if row[k] is not None else None for k in oxide_cols]
        except (ValueError, TypeError):
            continue
        if any(v is None or v <= 0 for v in vals): continue
        country = (str(row[country_idx]) or "").strip()
        if not country: continue
        by_country[country].append(vals)

    kept = {k:v for k,v in by_country.items() if len(v) >= MIN_N}
    print(f"[Tappe KIM1] clean rows: {sum(len(v) for v in by_country.values())}, "
          f"countries: {len(by_country)}, kept: {len(kept)}, samples kept: {sum(len(v) for v in kept.values())}")

    summary = []
    with (OUT / "tappe_kim1_by_country_barycenters.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Country_Region"] + OXIDES)
        for country, vals_list in sorted(kept.items()):
            closed = [[v/sum(vs) for v in vs] for vs in vals_list]
            bary = aitchison_barycenter(closed)
            short = country.split("(")[0].strip().replace(" ","_").replace(",","")[:30]
            w.writerow([short] + [f"{v:.10f}" for v in bary])
            summary.append({"country": country, "n": len(vals_list), "barycenter": bary})
            print(f"  {short:<32s}  n={len(vals_list)}")

    with (OUT / "tappe_kim1_by_country_summary.json").open("w") as f:
        json.dump({"scheme": "Tappe v2024 KIM1 by Country/Region",
                   "source": str(src.name), "min_samples": MIN_N,
                   "oxides": OXIDES, "countries": summary}, f, indent=2)
    return len(kept)


# ── (b) Qin clinopyroxene ──
def bin_qin_cpx(top_n=30):
    src = ROOT / "DATA" / "Geochemistry" / "2024-007_AVAW2Y_Qin_data.xlsx"
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb["Major elements"]

    # Header at L3, 28 cols. Find by name to be robust.
    rows_iter = ws.iter_rows(values_only=True)
    for _ in range(3): next(rows_iter)
    header = [str(c).strip() if c else "" for c in next(rows_iter)]

    OXIDE_NAMES = ["SIO2(WT%)","TIO2(WT%)","AL2O3(WT%)","CR2O3(WT%)",
                   "FEOT(WT%)","CAO(WT%)","MGO(WT%)","MNO(WT%)","NA2O(WT%)"]
    OUT_NAMES   = ["SiO2","TiO2","Al2O3","Cr2O3","FeO","CaO","MgO","MnO","Na2O"]
    oxide_idx = [header.index(n) for n in OXIDE_NAMES]
    loc_idx   = header.index("LOCATION")
    rock_idx  = header.index("ROCK NAME")

    by_loc = collections.defaultdict(list)
    by_loc_rock = collections.defaultdict(list)
    for row in rows_iter:
        if row is None: continue
        try:
            vals = [float(row[k]) if row[k] is not None else None for k in oxide_idx]
        except (ValueError, TypeError):
            continue
        if any(v is None or v <= 0 for v in vals): continue
        loc = (str(row[loc_idx]) or "").strip()
        if not loc: continue
        rock = (str(row[rock_idx]) or "").strip()
        by_loc[loc].append(vals)
        by_loc_rock[loc].append(rock)

    kept_all = {k:v for k,v in by_loc.items() if len(v) >= MIN_N}
    print(f"[Qin Cpx] clean rows: {sum(len(v) for v in by_loc.values())}, "
          f"locations: {len(by_loc)}, kept (n>={MIN_N}): {len(kept_all)}")

    # Take top-N by sample count for a manageable run
    top_locs = sorted(kept_all.items(), key=lambda kv: -len(kv[1]))[:top_n]
    print(f"  Using top {top_n} by sample count: total {sum(len(v) for _,v in top_locs)} samples")

    summary = []
    with (OUT / "qin_cpx_by_location_barycenters.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Location"] + OUT_NAMES)
        for loc, vals_list in top_locs:
            closed = [[v/sum(vs) for v in vs] for vs in vals_list]
            bary = aitchison_barycenter(closed)
            short = loc.split("/")[-1].strip().replace(" ","_").replace(",","")[:30]
            w.writerow([short] + [f"{v:.10f}" for v in bary])
            top_rock = collections.Counter(by_loc_rock[loc]).most_common(1)
            summary.append({
                "location": loc, "n": len(vals_list),
                "dominant_rock": top_rock[0][0] if top_rock else None,
                "barycenter": bary,
            })

    with (OUT / "qin_cpx_by_location_summary.json").open("w") as f:
        json.dump({"scheme": "Qin clinopyroxene by Location (top-N)",
                   "source": str(src.name), "top_n": top_n,
                   "oxides": OUT_NAMES, "locations": summary}, f, indent=2)
    return len(top_locs)


if __name__ == "__main__":
    print("=== Tappe v2024 KIM1 (kimberlite) ===")
    n_tappe = bin_tappe_kim1()
    print(f"\n=== Qin (intra-cratonic clinopyroxene, D=9) ===")
    n_qin = bin_qin_cpx(top_n=30)
    print(f"\nDone. Tappe T={n_tappe}, Qin T={n_qin}")
